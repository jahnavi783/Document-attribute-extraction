"""
File Writers - Generate normalized output files preserving original format.
"""

import io
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
)
from reportlab.lib.enums import TA_LEFT, TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


# ── PDF WRITER ────────────────────────────────────────────────────────────────

def write_pdf_keyvalue(records: list[dict], title: str = "Normalized Document") -> bytes:
    """Write key-value records to a styled PDF."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=2*cm, leftMargin=2*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  fontSize=16, spaceAfter=12, alignment=TA_CENTER,
                                  textColor=colors.HexColor("#1a237e"))
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.4*cm))

    # Build table data
    header_style = ParagraphStyle("hdr", parent=styles["Normal"],
                                   fontSize=10, fontName="Helvetica-Bold",
                                   textColor=colors.white)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                 fontSize=10, leading=14)

    table_data = [[Paragraph("Output Attribute", header_style),
                   Paragraph("Value", header_style)]]

    for rec in records:
        table_data.append([
            Paragraph(str(rec.get("attribute", "")), cell_style),
            Paragraph(str(rec.get("value", "")), cell_style),
        ])

    col_widths = [7.5*cm, 10.5*cm]
    t = Table(table_data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a237e")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.HexColor("#f5f5f5"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cccccc")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def write_pdf_tabular(dfs: list[pd.DataFrame], title: str = "Normalized Document") -> bytes:
    """Write tabular data to PDF."""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                             rightMargin=1*cm, leftMargin=1*cm,
                             topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story = []

    title_style = ParagraphStyle("title", parent=styles["Title"],
                                  fontSize=14, spaceAfter=10, alignment=TA_CENTER,
                                  textColor=colors.HexColor("#1a237e"))
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 0.3*cm))

    for i, df in enumerate(dfs):
        if df.empty:
            continue
        if len(dfs) > 1:
            story.append(Paragraph(f"Table {i+1}", styles["Heading2"]))

        col_style = ParagraphStyle("col", parent=styles["Normal"],
                                    fontSize=8, fontName="Helvetica-Bold",
                                    textColor=colors.white)
        cell_style = ParagraphStyle("cell", parent=styles["Normal"],
                                     fontSize=8, leading=11)

        # Headers
        table_data = [[Paragraph(str(c), col_style) for c in df.columns]]
        for _, row in df.iterrows():
            table_data.append([Paragraph(str(v), cell_style) for v in row])

        n_cols = len(df.columns)
        avail = 19 * cm
        col_w = [avail / n_cols] * n_cols

        t = Table(table_data, colWidths=col_w, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#283593")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.HexColor("#f5f5f5"), colors.white]),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 0.4*cm))

    doc.build(story)
    return buf.getvalue()


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────

def _header_style():
    return {
        "font": Font(bold=True, color="FFFFFF", size=11),
        "fill": PatternFill("solid", start_color="1A237E"),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
    }

def _apply(cell, **kwargs):
    for k, v in kwargs.items():
        setattr(cell, k, v)


def write_excel_keyvalue(records: list[dict], sheet_name: str = "Normalized") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr = _header_style()
    for col, title in [(1, "Output Attribute"), (2, "Value")]:
        c = ws.cell(row=1, column=col, value=title)
        _apply(c, **hdr)
        ws.row_dimensions[1].height = 22

    thin = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    for i, rec in enumerate(records, start=2):
        fill_color = "F5F5F5" if i % 2 == 0 else "FFFFFF"
        fill = PatternFill("solid", start_color=fill_color)
        for col, key in [(1, "attribute"), (2, "value")]:
            c = ws.cell(row=i, column=col, value=rec.get(key, ""))
            c.border = thin
            c.fill = fill
            c.alignment = Alignment(vertical="top", wrap_text=True)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def write_excel_tabular(sheets: dict[str, pd.DataFrame]) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        hdr = _header_style()
        thin = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC"),
        )

        for col_idx, col_name in enumerate(df.columns, start=1):
            c = ws.cell(row=1, column=col_idx, value=col_name)
            _apply(c, **hdr)

        for r_idx, (_, row) in enumerate(df.iterrows(), start=2):
            fill_color = "F5F5F5" if r_idx % 2 == 0 else "FFFFFF"
            fill = PatternFill("solid", start_color=fill_color)
            for col_idx, val in enumerate(row, start=1):
                c = ws.cell(row=r_idx, column=col_idx, value=val)
                c.border = thin
                c.fill = fill
                c.alignment = Alignment(vertical="top", wrap_text=True)

        for col_idx, col in enumerate(df.columns, start=1):
            max_len = max(
                len(str(col)),
                *(len(str(v)) for v in df.iloc[:, col_idx - 1]) if len(df) else [0]
            )
            ws.column_dimensions[ws.cell(1, col_idx).column_letter].width = min(max_len + 4, 45)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── CSV WRITER ────────────────────────────────────────────────────────────────

def write_csv(df: pd.DataFrame) -> bytes:
    buf = io.StringIO()
    df.to_csv(buf, index=False)
    return buf.getvalue().encode("utf-8")
